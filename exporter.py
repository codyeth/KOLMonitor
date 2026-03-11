"""
exporter.py — Xuất kết quả ra Excel (.xlsx) và CSV (.csv)
"""

import csv
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


HEADER_COLOR = "1DA1F2"   # màu X/Twitter
ROW_ALT_COLOR = "F8F9FA"  # xám nhạt xen kẽ


def _fmt_date(created_at_str: str) -> str:
    """Chuyển chuỗi twikit sang DD/MM/YYYY HH:MM"""
    try:
        dt = datetime.strptime(created_at_str, "%a %b %d %H:%M:%S %z %Y")
        return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return created_at_str


def _source_label(source: str) -> str:
    return "KOL mới phát hiện" if source == "discovered" else "KOL đã theo dõi"


def export(
    kol_hits: list[dict],
    discovered: list[dict],
    keywords: list[str],
    mode: str,
    since_hours: int,
    min_views: int,
    output_dir: Path,
    output_name: str = None,
    csv_only: bool = False,
    slim: bool = False,
    append: bool = False,
) -> tuple[Path, Path]:
    """Tạo file Excel và CSV, trả về (excel_path, csv_path)."""
    output_dir.mkdir(parents=True, exist_ok=True)

    now = datetime.now(timezone.utc)
    date_str = now.strftime("%Y-%m-%d_%H%M")
    prefix = output_name if output_name else f"kol_report_{date_str}"
    xlsx_path = output_dir / f"{prefix}.xlsx"
    csv_path = output_dir / f"{prefix}.csv"

    all_tweets = []
    for t in kol_hits:
        t = dict(t)
        t.setdefault("source", "known_kol")
        all_tweets.append(t)
    for t in discovered:
        t = dict(t)
        t.setdefault("source", "discovered")
        all_tweets.append(t)

    if not csv_only:
        _write_excel(xlsx_path, all_tweets, kol_hits, discovered, keywords, mode, since_hours, min_views, now)
    else:
        xlsx_path = None
    _write_csv(csv_path, all_tweets, slim=slim, append=append)

    return xlsx_path, csv_path


# ── Excel ──────────────────────────────────────────────────────────────────────

def _write_excel(
    path: Path,
    all_tweets: list[dict],
    kol_hits: list[dict],
    discovered: list[dict],
    keywords: list[str],
    mode: str,
    since_hours: int,
    min_views: int,
    now: datetime,
):
    wb = Workbook()

    # Sheet 1: bài đăng
    ws1 = wb.active
    ws1.title = "Bài Đăng KOL"
    _fill_posts_sheet(ws1, all_tweets)

    # Sheet 2: tóm tắt
    ws2 = wb.create_sheet("Tóm Tắt")
    _fill_summary_sheet(ws2, all_tweets, kol_hits, discovered, keywords, mode, since_hours, min_views, now)

    wb.save(path)


def _header_style():
    fill = PatternFill("solid", fgColor=HEADER_COLOR)
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center", vertical="center")
    return fill, font, align


def _fill_posts_sheet(ws, tweets: list[dict]):
    headers = [
        "STT", "Ngày đăng", "Username", "Tên KOL", "Followers",
        "Lượt xem", "Likes", "Retweets", "Replies",
        "Nội dung", "Link bài viết", "Link X", "Loại", "Từ khóa khớp",
    ]

    fill, font, align = _header_style()
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = align

    ws.freeze_panes = "A2"

    alt_fill = PatternFill("solid", fgColor=ROW_ALT_COLOR)

    for row_idx, t in enumerate(tweets, 2):
        is_alt = (row_idx % 2 == 0)
        row_fill = alt_fill if is_alt else None

        def _cell(col, value):
            c = ws.cell(row=row_idx, column=col, value=value)
            if row_fill:
                c.fill = row_fill
            return c

        _cell(1, row_idx - 1)                             # STT
        _cell(2, _fmt_date(t.get("created_at", "")))      # Ngày đăng
        _cell(3, f"@{t.get('username', '')}")              # Username
        _cell(4, t.get("name", ""))                        # Tên KOL
        _cell(5, t.get("followers", 0))                    # Followers
        _cell(6, t.get("views", 0))                        # Lượt xem
        _cell(7, t.get("likes", 0))                        # Likes
        _cell(8, t.get("retweets", 0))                     # Retweets
        _cell(9, t.get("replies", 0))                      # Replies
        _cell(10, t.get("text", "")[:150])                 # Nội dung

        # Cột K: hyperlink
        url = t.get("url", "")
        link_cell = ws.cell(row=row_idx, column=11, value=url)
        if url:
            link_cell.hyperlink = url
            link_cell.font = Font(color=HEADER_COLOR, underline="single")
        if row_fill:
            link_cell.fill = row_fill

        # Cột L: Link X (profile)
        profile_url = t.get("profile_url", "")
        profile_cell = ws.cell(row=row_idx, column=12, value=profile_url)
        if profile_url:
            profile_cell.hyperlink = profile_url
            profile_cell.font = Font(color=HEADER_COLOR, underline="single")
        if row_fill:
            profile_cell.fill = row_fill

        _cell(13, _source_label(t.get("source", "")))     # Loại
        _cell(14, t.get("matched_keywords", ""))           # Từ khóa khớp

        # Số định dạng có dấu phẩy
        ws.cell(row=row_idx, column=5).number_format = "#,##0"
        ws.cell(row=row_idx, column=6).number_format = "#,##0"
        ws.cell(row=row_idx, column=7).number_format = "#,##0"
        ws.cell(row=row_idx, column=8).number_format = "#,##0"
        ws.cell(row=row_idx, column=9).number_format = "#,##0"

    # Auto-fit column width
    col_widths = [6, 18, 20, 25, 12, 12, 8, 10, 8, 60, 45, 45, 20, 25]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _fill_summary_sheet(ws, all_tweets, kol_hits, discovered, keywords, mode, since_hours, min_views, now):
    fill, font, align = _header_style()

    # Tiêu đề
    title_cell = ws.cell(row=1, column=1, value="KOL Monitor — Báo Cáo Tóm Tắt")
    title_cell.font = Font(color="FFFFFF", bold=True, size=13)
    title_cell.fill = fill
    ws.merge_cells("A1:B1")
    ws.cell(row=1, column=2).fill = fill

    total_views = sum(t.get("views", 0) for t in all_tweets)
    top_tweet = max(all_tweets, key=lambda x: x.get("views", 0), default=None)
    top_str = f"@{top_tweet['username']} — {top_tweet['views']:,} views" if top_tweet else "—"

    mode_label = "daily (24h gần nhất)" if mode == "daily" else f"on-demand ({since_hours}h gần nhất)"

    rows = [
        ("Thời gian chạy", now.strftime("%d/%m/%Y %H:%M UTC")),
        ("Chế độ", mode_label),
        ("Khoảng thời gian", f"{since_hours}h"),
        ("Từ khóa", ", ".join(keywords)),
        ("Lượt xem tối thiểu", f"{min_views:,}"),
        ("", ""),
        ("Tổng bài từ KOL đã theo dõi", len(kol_hits)),
        ("Tổng bài từ KOL mới phát hiện", len(discovered)),
        ("Tổng cộng", len(all_tweets)),
        ("Tổng lượt xem", f"{total_views:,}"),
        ("Bài xem nhiều nhất", top_str),
    ]

    label_font = Font(bold=True)
    for r_idx, (label, value) in enumerate(rows, 2):
        lc = ws.cell(row=r_idx, column=1, value=label)
        vc = ws.cell(row=r_idx, column=2, value=value)
        if label:
            lc.font = label_font

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 40


# ── CSV ────────────────────────────────────────────────────────────────────────

def _write_csv(path: Path, tweets: list[dict], slim: bool = False, append: bool = False):
    if slim:
        fieldnames = ["link_bai_viet", "link_x"]
    else:
        fieldnames = [
            "stt", "ngay_dang", "username", "ten_kol", "followers",
            "luot_xem", "likes", "retweets", "replies",
            "noi_dung_150ky", "link_bai_viet", "link_x", "loai", "tu_khoa_khop",
        ]

    file_exists = path.exists() and append
    mode = "a" if file_exists else "w"
    with open(path, mode, newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if not file_exists:
            writer.writeheader()
        for idx, t in enumerate(tweets, 1):
            if slim:
                writer.writerow({
                    "link_bai_viet": t.get("url", ""),
                    "link_x": t.get("profile_url", ""),
                })
            else:
                writer.writerow({
                    "stt": idx,
                    "ngay_dang": _fmt_date(t.get("created_at", "")),
                    "username": t.get("username", ""),
                    "ten_kol": t.get("name", ""),
                    "followers": t.get("followers", 0),
                    "luot_xem": t.get("views", 0),
                    "likes": t.get("likes", 0),
                    "retweets": t.get("retweets", 0),
                    "replies": t.get("replies", 0),
                    "noi_dung_150ky": t.get("text", "")[:150],
                    "link_bai_viet": t.get("url", ""),
                    "link_x": t.get("profile_url", ""),
                    "loai": _source_label(t.get("source", "")),
                    "tu_khoa_khop": t.get("matched_keywords", ""),
                })

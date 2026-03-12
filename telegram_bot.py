"""
telegram_bot.py — Telegram bot interface cho KOL Monitor

Commands:
  /start         — Chào mừng + hướng dẫn
  /help          — Hướng dẫn sử dụng
  /myid          — Hiển thị Telegram user ID của bạn
  /adduser <id>  — Thêm user (admin only)
  /removeuser <id> — Xoá user (admin only)
  /listusers     — Danh sách user được phép (admin only)
  /monitor       — Kiểm tra bài mới từ KOL đã có trong data/kols.json
  /scan          — Quét X tìm KOL mới đang đăng về dự án (24h, view > 1k)
  /scan KW1 KW2  — Quét với từ khóa tuỳ chọn

Gửi danh sách KOL ad-hoc (URL hoặc username, mỗi dòng 1 tài khoản):
  https://x.com/user1
  https://x.com/user2
  @user3
  user4

Auto scan: chạy tự động lúc 09:00 Asia/Ho_Chi_Minh, gửi kết quả cho admin.
"""

import asyncio
import json
import logging
import os
import re
import tempfile
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from telegram import BotCommand, Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters,
)

from config import (
    CONFIG,
    SCAN_HOURS,
    SCAN_MAX_RESULTS,
    SCAN_MIN_VIEWS,
    TELEGRAM_ADMIN_ID,
    TELEGRAM_ALLOWED_IDS,
    TELEGRAM_TOKEN,
)
from monitor_core import run_monitor

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# File lưu danh sách allowed IDs (để persist qua restart)
ALLOWED_IDS_FILE = Path(__file__).parent / "data" / "allowed_users.json"


# ── Quản lý allowed users ────────────────────────────────────────────────────

def _load_allowed_ids() -> set[int]:
    ids = set(TELEGRAM_ALLOWED_IDS)
    if TELEGRAM_ADMIN_ID:
        ids.add(TELEGRAM_ADMIN_ID)
    if ALLOWED_IDS_FILE.exists():
        try:
            ids.update(json.loads(ALLOWED_IDS_FILE.read_text()))
        except Exception:
            pass
    return ids


def _save_allowed_ids(ids: set[int]):
    ALLOWED_IDS_FILE.parent.mkdir(parents=True, exist_ok=True)
    ALLOWED_IDS_FILE.write_text(json.dumps(sorted(ids)))


allowed_ids: set[int] = _load_allowed_ids()

# ── Twitter fetcher singleton (tái dùng session, không login lại mỗi lần) ────

_fetcher = None


async def _get_fetcher():
    global _fetcher
    if _fetcher is None:
        from fetcher import TwitterFetcher
        _fetcher = TwitterFetcher()
        await _fetcher.login()
    return _fetcher


def is_allowed(user_id: int) -> bool:
    return user_id in allowed_ids


def is_admin(user_id: int) -> bool:
    return user_id == TELEGRAM_ADMIN_ID


# ── Parse usernames từ message ───────────────────────────────────────────────

def parse_usernames(text: str) -> list[str]:
    """
    Nhận text chứa URLs, @mentions hoặc usernames thuần.
    Trả về list username (không có @, không trùng, giữ thứ tự).
    """
    usernames = []
    seen = set()

    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue

        # x.com/username hoặc twitter.com/username
        m = re.search(r"(?:x\.com|twitter\.com)/([A-Za-z0-9_]{1,50})", line)
        if m:
            username = m.group(1)
        elif line.startswith("@"):
            username = line.lstrip("@")
        elif re.fullmatch(r"[A-Za-z0-9_]{1,50}", line):
            username = line
        else:
            continue

        lower = username.lower()
        if lower not in seen:
            seen.add(lower)
            usernames.append(username)

    return usernames


# ── Helpers ──────────────────────────────────────────────────────────────────

def _main_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[
        InlineKeyboardButton("📋 Monitor", callback_data="menu_monitor"),
        InlineKeyboardButton("🔍 Scan", callback_data="cmd_scan"),
        InlineKeyboardButton("❓ Help", callback_data="menu_help"),
    ]])


def _monitor_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("🐦 Monitor X", callback_data="cmd_monitor_x"),
            InlineKeyboardButton("💬 Monitor TG", callback_data="cmd_monitor_tg"),
        ],
        [InlineKeyboardButton("🔙 Quay lại", callback_data="menu_main")],
    ])


# ── Handlers ─────────────────────────────────────────────────────────────────

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_allowed(user_id):
        await update.message.reply_text(
            f"⛔ Bạn chưa được cấp quyền.\n\n"
            f"🪪 ID của bạn: <code>{user_id}</code>\n\n"
            f"Gửi ID này cho admin để được thêm vào.",
            parse_mode="HTML",
        )
        return

    await update.message.reply_text(
        "👋 <b>Chào mừng đến với KOL Monitor Bot!</b>\n\n"
        "Bot giúp bạn theo dõi KOL trên X/Twitter và Telegram đang đăng về dự án.\n\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "📋 <b>/monitor_x</b> — Kiểm tra KOL X/Twitter\n"
        "💬 <b>/monitor_tg</b> — Kiểm tra KOL Telegram\n"
        "🔍 <b>/scan</b> — Tìm KOL mới đăng về dự án\n"
        "❓ <b>/help</b> — Hướng dẫn chi tiết\n"
        "━━━━━━━━━━━━━━━━━━━━\n\n"
        "👇 Nhấn để chọn tính năng:",
        parse_mode="HTML",
        reply_markup=_main_keyboard(),
    )


async def cmd_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update.effective_user.id):
        return

    await update.message.reply_text(
        "❓ <b>Hướng dẫn sử dụng KOL Monitor Bot</b>\n\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "📋 <b>MONITOR — Theo dõi KOL có sẵn</b>\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "Dùng để kiểm tra bài mới từ danh sách KOL bạn cung cấp.\n\n"
        "📌 <b>Cách dùng:</b>\n"
        "Gõ <code>/monitor</code> kèm danh sách username:\n\n"
        "<code>/monitor\n"
        "https://x.com/user1\n"
        "@user2\n"
        "user3</code>\n\n"
        "⚙️ <b>Điều kiện lọc:</b>\n"
        "• Bài trong 72h gần nhất\n"
        "• Chứa từ khóa dự án (NEXI, Nexira, DAEP...)\n"
        "• Kết quả xuất ra file CSV\n\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "🔍 <b>SCAN — Tìm KOL mới trên X</b>\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "Dùng để tìm các tài khoản đang đăng về dự án mà chưa có trong danh sách.\n\n"
        "📌 <b>Cách dùng:</b>\n"
        "• <code>/scan</code> — Dùng từ khóa mặc định\n"
        "• <code>/scan NEXI</code> — Scan từ khóa tùy chọn\n"
        "• <code>/scan NEXI Nexira DAEP</code> — Nhiều từ khóa\n\n"
        "⚙️ <b>Điều kiện lọc:</b>\n"
        "• Bài trong 24h gần nhất\n"
        "• View trên 1,000\n"
        "• Kết quả gửi message + file Excel\n"
        "• Tự động chạy lúc 9:00 sáng mỗi ngày\n\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "🛠 <b>Lệnh khác</b>\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "• <code>/myid</code> — Xem Telegram ID của bạn\n"
        "• <code>/adduser ID</code> — Thêm user (admin)\n"
        "• <code>/removeuser ID</code> — Xóa user (admin)\n"
        "• <code>/listusers</code> — Danh sách user (admin)\n\n"
        "💡 <b>Gửi link trực tiếp:</b>\n"
        "Bạn cũng có thể gửi danh sách username vào chat mà không cần gõ lệnh:\n"
        "<code>https://x.com/user1\n@user2\nuser3</code>",
        parse_mode="HTML",
        disable_web_page_preview=True,
    )


async def cmd_button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý khi user nhấn button inline."""
    query = update.callback_query
    await query.answer()

    if query.data == "menu_main":
        await query.message.reply_text(
            "👇 Chọn tính năng:",
            reply_markup=_main_keyboard(),
        )

    elif query.data == "menu_monitor":
        await query.message.reply_text(
            "📋 <b>Chọn nền tảng muốn theo dõi:</b>",
            parse_mode="HTML",
            reply_markup=_monitor_keyboard(),
        )

    elif query.data == "cmd_monitor_x":
        context.user_data["mode"] = "monitor_x"
        await query.message.reply_text(
            "🐦 <b>Monitor X/Twitter</b>\n\n"
            "Gửi danh sách KOL cần kiểm tra (mỗi dòng 1 tài khoản):\n\n"
            "<code>https://x.com/user1\n"
            "@user2\n"
            "user3</code>\n\n"
            "⚙️ Lọc: 72h gần nhất · Từ khóa dự án\n"
            "📤 Kết quả: file CSV",
            parse_mode="HTML",
            reply_markup=_monitor_keyboard(),
        )

    elif query.data == "cmd_monitor_tg":
        context.user_data["mode"] = "monitor_tg"
        await query.message.reply_text(
            "💬 <b>Monitor Telegram</b>\n\n"
            "Gửi danh sách channel/group cần quét (mỗi dòng 1 link):\n\n"
            "<code>https://t.me/channel1\n"
            "https://t.me/group2\n"
            "@channelname</code>\n\n"
            "⚙️ Lọc: 24h gần nhất · Từ khóa dự án\n"
            "📤 Kết quả: file CSV",
            parse_mode="HTML",
            reply_markup=_monitor_keyboard(),
        )

    elif query.data == "cmd_scan":
        if not is_allowed(query.from_user.id):
            return
        keywords = CONFIG["keywords"]
        kw_str = ", ".join(keywords)
        status_msg = await query.message.reply_text(
            f"⏳ Đang quét X/Twitter...\n🔑 Keywords: <b>{kw_str}</b>\n\nVui lòng chờ ~30 giây",
            parse_mode="HTML",
        )

        last_text = [""]

        async def progress_cb(msg: str):
            if msg != last_text[0]:
                last_text[0] = msg
                try:
                    await status_msg.edit_text(msg, parse_mode="HTML")
                except Exception:
                    pass

        try:
            tweets = await _scan_kol_tweets(keywords, progress_cb=progress_cb)
        except Exception as e:
            await status_msg.edit_text(f"❌ Lỗi khi quét:\n<code>{e}</code>", parse_mode="HTML")
            return

        await status_msg.delete()

        async def send_msg(text):
            await query.message.reply_text(text, parse_mode="HTML", disable_web_page_preview=True)

        async def send_doc(f, filename, caption):
            await query.message.reply_document(document=f, filename=filename, caption=caption)

        await _send_scan_results(send_msg, send_doc, tweets, keywords)
        await query.message.reply_text("Bạn muốn làm gì tiếp theo?", reply_markup=_main_keyboard())

    elif query.data == "menu_help":
        await query.message.reply_text(
            "❓ <b>Hướng dẫn sử dụng KOL Monitor Bot</b>\n\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            "📋 <b>MONITOR X</b>\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            "Theo dõi danh sách KOL trên X/Twitter.\n"
            "Gửi username hoặc link → bot kiểm tra bài đăng 72h gần nhất có từ khóa dự án.\n\n"
            "Lệnh: <code>/monitor_x</code>\n"
            "Gửi trực tiếp: <code>https://x.com/username</code>\n\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            "💬 <b>MONITOR TG</b>\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            "Quét channel/group Telegram tìm bài đăng có từ khóa dự án trong 24h qua.\n\n"
            "Lệnh: <code>/monitor_tg</code>\n"
            "Gửi trực tiếp: <code>https://t.me/channel</code>\n\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            "🔍 <b>SCAN X</b>\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            "Tìm KOL mới đang đăng về dự án. Không cần danh sách — bot tự tìm.\n\n"
            "Lệnh: <code>/scan</code>\n"
            "Tùy chỉnh: <code>/scan NEXI Nexira</code>\n\n"
            "⚙️ Điều kiện: 24h · view > 1,000\n"
            "⏱ Tự động: 9:00 sáng mỗi ngày\n\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            "📤 Output: CSV (Monitor) · Excel (Scan)",
            parse_mode="HTML",
            reply_markup=_main_keyboard(),
        )

    else:
        # Button cũ (từ trước khi redeploy) — hướng dẫn dùng lại /start
        await query.message.reply_text(
            "⚠️ Button này đã hết hạn do bot vừa cập nhật.\n\n"
            "Gõ /start để lấy menu mới.",
            reply_markup=_main_keyboard(),
        )


async def cmd_myid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    await update.message.reply_text(f"🪪 Telegram ID của bạn: `{uid}`", parse_mode="Markdown")


async def cmd_adduser(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Chỉ admin mới dùng được lệnh này.")
        return
    if not context.args:
        await update.message.reply_text("Cách dùng: /adduser <telegram_id>")
        return
    try:
        new_id = int(context.args[0])
        allowed_ids.add(new_id)
        _save_allowed_ids(allowed_ids)
        await update.message.reply_text(f"✅ Đã thêm user `{new_id}`.", parse_mode="Markdown")
    except ValueError:
        await update.message.reply_text("❌ ID không hợp lệ.")


async def cmd_removeuser(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Chỉ admin mới dùng được lệnh này.")
        return
    if not context.args:
        await update.message.reply_text("Cách dùng: /removeuser <telegram_id>")
        return
    try:
        rem_id = int(context.args[0])
        allowed_ids.discard(rem_id)
        _save_allowed_ids(allowed_ids)
        await update.message.reply_text(f"✅ Đã xoá user `{rem_id}`.", parse_mode="Markdown")
    except ValueError:
        await update.message.reply_text("❌ ID không hợp lệ.")


async def cmd_listusers(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Chỉ admin mới dùng được lệnh này.")
        return
    if not allowed_ids:
        await update.message.reply_text("Chưa có user nào.")
        return
    lines = [f"• `{uid}`{'  ← admin' if uid == TELEGRAM_ADMIN_ID else ''}" for uid in sorted(allowed_ids)]
    await update.message.reply_text(
        "👥 *Danh sách user được phép:*\n" + "\n".join(lines),
        parse_mode="Markdown",
    )


async def cmd_monitor(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Hiển thị sub-menu chọn Monitor X hoặc Monitor TG."""
    if not is_allowed(update.effective_user.id):
        return
    await update.message.reply_text(
        "📋 <b>Chọn loại Monitor:</b>",
        parse_mode="HTML",
        reply_markup=_monitor_keyboard(),
    )


async def cmd_monitor_x(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update.effective_user.id):
        return

    full_text = update.message.text or ""
    body = re.sub(r"^/monitor_x\S*\s*", "", full_text, count=1).strip()
    usernames = parse_usernames(body)

    if not usernames:
        await update.message.reply_text(
            "📋 *Cách dùng /monitor:*\n\n"
            "Gửi lệnh kèm danh sách KOL:\n"
            "```\n/monitor_x\nhttps://x.com/user1\n@user2\nuser3\n```",
            parse_mode="Markdown",
            reply_markup=_monitor_keyboard(),
        )
        return

    status_msg = await update.message.reply_text(
        f"⏳ Bắt đầu kiểm tra {len(usernames)} KOL...\n\n"
        + "\n".join(f"• @{u}" for u in usernames),
    )

    progress_lines: list[str] = []

    async def on_progress(msg: str):
        progress_lines.append(msg)
        if len(progress_lines) % 3 == 0 or "✗" in msg:
            try:
                await status_msg.edit_text(
                    "⏳ Đang kiểm tra...\n\n" + "\n".join(progress_lines[-15:])
                )
            except Exception:
                pass

    try:
        _, csv_path, kol_hits, _ = await run_monitor(
            usernames=usernames,
            on_progress=on_progress,
        )
    except Exception as e:
        await status_msg.edit_text(f"❌ Lỗi:\n`{e}`", parse_mode="Markdown")
        return

    if not kol_hits:
        await status_msg.edit_text(
            f"✅ Đã kiểm tra {len(usernames)} KOL\n\n"
            "ℹ️ Không có bài viết nào khớp từ khoá trong 72 giờ gần nhất.",
            reply_markup=_main_keyboard(),
        )
        return

    total_views = sum(t.get("views", 0) for t in kol_hits)
    await status_msg.edit_text(
        f"✅ Hoàn tất!\n\n"
        f"📊 {len(kol_hits)} bài viết khớp\n"
        f"👁 {total_views:,} lượt xem tổng\n\n"
        "Đang gửi file CSV..."
    )

    with open(csv_path, "rb") as f:
        await update.message.reply_document(
            document=f,
            filename=csv_path.name,
            caption=f"📄 {len(kol_hits)} bài | {total_views:,} views | {len(usernames)} KOL",
        )
    await update.message.reply_text("Bạn muốn làm gì tiếp theo?", reply_markup=_main_keyboard())


async def cmd_monitor_tg(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update.effective_user.id):
        return

    full_text = update.message.text or ""
    body = re.sub(r"^/monitor_tg\S*\s*", "", full_text, count=1).strip()

    # Parse Telegram channel usernames (t.me/xxx, @xxx, hoặc username thuần)
    tg_usernames = []
    seen = set()
    for line in body.splitlines():
        line = line.strip()
        if not line:
            continue
        m = re.search(r"t\.me/([A-Za-z0-9_]{3,})", line)
        if m:
            uname = m.group(1)
        elif line.startswith("@"):
            uname = line.lstrip("@")
        elif re.fullmatch(r"[A-Za-z0-9_]{3,}", line):
            uname = line
        else:
            continue
        low = uname.lower()
        if low not in seen:
            seen.add(low)
            tg_usernames.append(uname)

    if not tg_usernames:
        await update.message.reply_text(
            "💬 <b>Cách dùng /monitor_tg:</b>\n\n"
            "<code>/monitor_tg\n"
            "https://t.me/channel1\n"
            "@channel2\n"
            "channel3</code>",
            parse_mode="HTML",
            reply_markup=_monitor_keyboard(),
        )
        return

    await update.message.reply_text(
        "💬 <b>Monitor TG</b>\n\n"
        f"Tính năng đang phát triển.\n"
        f"Các channel được ghi nhận:\n"
        + "\n".join(f"• @{u}" for u in tg_usernames)
        + "\n\n<i>Vui lòng theo dõi cập nhật sau.</i>",
        parse_mode="HTML",
        reply_markup=_main_keyboard(),
    )


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    if not is_allowed(user_id):
        await update.message.reply_text(
            f"⛔ Bạn chưa được cấp quyền.\nID của bạn: `{user_id}`",
            parse_mode="Markdown",
        )
        return

    mode = context.user_data.get("mode", "monitor_x")
    text = update.message.text or ""

    # ── Mode: Monitor TG ─────────────────────────────────────────────────────
    if mode == "monitor_tg":
        tg_usernames = []
        seen = set()
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            m = re.search(r"t\.me/([A-Za-z0-9_]{3,})", line)
            if m:
                uname = m.group(1)
            elif line.startswith("@"):
                uname = line.lstrip("@")
            elif re.fullmatch(r"[A-Za-z0-9_]{3,}", line):
                uname = line
            else:
                continue
            low = uname.lower()
            if low not in seen:
                seen.add(low)
                tg_usernames.append(uname)

        if not tg_usernames:
            await update.message.reply_text(
                "❓ Không nhận ra channel/group nào.\n\n"
                "Gửi danh sách theo dạng:\n"
                "• https://t.me/channel\n"
                "• @channel\n"
                "• channelname",
                reply_markup=_monitor_keyboard(),
            )
            return

        context.user_data.pop("mode", None)
        status_msg = await update.message.reply_text(
            f"⏳ Đang quét {len(tg_usernames)} channel Telegram...\n\n"
            + "\n".join(f"• @{u}" for u in tg_usernames),
        )

        try:
            from monitors.tg_fetcher import fetch_tg_channels
            messages, errors = await fetch_tg_channels(
                tg_usernames,
                keywords=CONFIG["keywords"],
                hours=24,
            )
        except RuntimeError as e:
            await status_msg.edit_text(
                f"❌ Lỗi cấu hình Monitor TG:\n<code>{e}</code>",
                parse_mode="HTML",
                reply_markup=_main_keyboard(),
            )
            return
        except Exception as e:
            await status_msg.edit_text(
                f"❌ Lỗi khi quét Telegram:\n<code>{e}</code>",
                parse_mode="HTML",
                reply_markup=_main_keyboard(),
            )
            return

        if not messages:
            err_note = ("\n\n⚠️ Lỗi channel:\n" + "\n".join(f"• {e}" for e in errors[:3])) if errors else ""
            await status_msg.edit_text(
                f"✅ Đã quét {len(tg_usernames)} channel\n\n"
                f"ℹ️ Không có tin nhắn nào khớp từ khoá trong 24 giờ gần nhất.{err_note}",
                reply_markup=_main_keyboard(),
            )
            return

        # Tạo CSV
        import csv
        import tempfile
        today = datetime.now().strftime("%Y-%m-%d")
        csv_path = os.path.join(tempfile.gettempdir(), f"tg_monitor_{today}.csv")
        fieldnames = [
            "stt", "ngay_dang", "username", "ten_kol", "followers",
            "luot_xem", "likes", "forwards", "replies",
            "noi_dung_150ky", "link_bai_viet", "tu_khoa_khop",
        ]
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for i, m in enumerate(messages, 1):
                try:
                    dt = datetime.strptime(m["created_at"], "%a %b %d %H:%M:%S %z %Y")
                    date_str = dt.strftime("%d/%m/%Y %H:%M")
                except Exception:
                    date_str = m.get("created_at", "")
                writer.writerow({
                    "stt": i,
                    "ngay_dang": date_str,
                    "username": m["username"],
                    "ten_kol": m["name"],
                    "followers": m["followers"],
                    "luot_xem": m["views"],
                    "likes": m["likes"],
                    "forwards": m["forwards"],
                    "replies": m["replies"],
                    "noi_dung_150ky": m["text"],
                    "link_bai_viet": m["url"],
                    "tu_khoa_khop": m["matched_keywords"],
                })

        total_views = sum(m.get("views", 0) for m in messages)
        err_note = (f"\n⚠️ {len(errors)} channel lỗi" if errors else "")
        await status_msg.edit_text(
            f"✅ Hoàn tất Monitor TG!\n\n"
            f"💬 {len(messages)} tin nhắn khớp\n"
            f"👁 {total_views:,} lượt xem tổng\n"
            f"📡 {len(tg_usernames)} channel{err_note}\n\n"
            "Đang gửi file CSV..."
        )

        with open(csv_path, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename=os.path.basename(csv_path),
                caption=f"💬 {len(messages)} tin | {total_views:,} views | {len(tg_usernames)} channel",
            )
        os.remove(csv_path)
        await update.message.reply_text("Bạn muốn làm gì tiếp theo?", reply_markup=_main_keyboard())
        return

    # ── Mode: Monitor X (default) ─────────────────────────────────────────────
    usernames = parse_usernames(text)

    if not usernames:
        await update.message.reply_text(
            "❓ Không nhận ra tài khoản nào.\n\n"
            "Gửi danh sách theo dạng:\n"
            "• https://x.com/username\n"
            "• @username\n"
            "• username"
        )
        return

    context.user_data.pop("mode", None)
    status_msg = await update.message.reply_text(
        f"⏳ Bắt đầu kiểm tra {len(usernames)} KOL...\n\n"
        + "\n".join(f"• @{u}" for u in usernames),
    )

    progress_lines: list[str] = []

    async def on_progress(msg: str):
        progress_lines.append(msg)
        if len(progress_lines) % 3 == 0 or "✗" in msg:
            try:
                await status_msg.edit_text(
                    "⏳ Đang kiểm tra...\n\n" + "\n".join(progress_lines[-15:]),
                )
            except Exception:
                pass

    try:
        _, csv_path, kol_hits, _ = await run_monitor(
            usernames=usernames,
            on_progress=on_progress,
        )
    except Exception as e:
        await status_msg.edit_text(f"❌ Lỗi khi chạy monitor:\n`{e}`", parse_mode="Markdown")
        return

    if not kol_hits:
        await status_msg.edit_text(
            f"✅ Đã kiểm tra {len(usernames)} KOL\n\n"
            "ℹ️ Không có bài viết nào khớp từ khoá trong 72 giờ gần nhất.",
            reply_markup=_main_keyboard(),
        )
        return

    total_views = sum(t.get("views", 0) for t in kol_hits)
    await status_msg.edit_text(
        f"✅ Hoàn tất!\n\n"
        f"📊 {len(kol_hits)} bài viết khớp\n"
        f"👁 {total_views:,} lượt xem tổng\n\n"
        f"Đang gửi file CSV...",
    )

    with open(csv_path, "rb") as f:
        await update.message.reply_document(
            document=f,
            filename=csv_path.name,
            caption=f"📄 {len(kol_hits)} bài | {total_views:,} views | {len(usernames)} KOL",
        )
    await update.message.reply_text("Bạn muốn làm gì tiếp theo?", reply_markup=_main_keyboard())


# ── Scan KOL theo từ khóa ────────────────────────────────────────────────────

async def _scan_kol_tweets(
    keywords: list[str],
    progress_cb=None,
) -> list[dict]:
    """
    Tìm tweet 24h gần nhất chứa keyword và views >= SCAN_MIN_VIEWS.
    progress_cb(msg): gửi cập nhật tiến độ (tuỳ chọn).
    """
    from filter import TweetFilter

    since_dt = datetime.now(timezone.utc) - timedelta(hours=SCAN_HOURS)
    tweet_filter = TweetFilter(keywords=keywords, min_views=SCAN_MIN_VIEWS, since=since_dt)

    fetcher = await _get_fetcher()
    seen_ids: set = set()
    raw: list[dict] = []

    for kw_idx, keyword in enumerate(keywords, 1):
        if progress_cb:
            await progress_cb(
                f"🔍 [{kw_idx}/{len(keywords)}] Đang quét: <b>{keyword}</b>..."
            )
        try:
            page_count = [0]

            async def on_page(_new_count, total, _kw=keyword, _ki=kw_idx):
                page_count[0] += 1
                if progress_cb:
                    await progress_cb(
                        f"🔍 [{_ki}/{len(keywords)}] <b>{_kw}</b>\n"
                        f"   Trang {page_count[0]} — thu thập {total} bài..."
                    )

            tweets = await fetcher.search_tweets_deep(
                keyword,
                max_results=SCAN_MAX_RESULTS,
                page_delay=4.0,
                on_page=on_page,
            )
            new = 0
            for t in tweets:
                if t["id"] not in seen_ids:
                    seen_ids.add(t["id"])
                    raw.append(t)
                    new += 1

            if progress_cb:
                await progress_cb(
                    f"✅ [{kw_idx}/{len(keywords)}] <b>{keyword}</b> — {new} bài mới"
                )

        except Exception as e:
            logger.error(f"Scan keyword '{keyword}' lỗi: {e}")
            if progress_cb:
                await progress_cb(f"⚠️ [{kw_idx}/{len(keywords)}] <b>{keyword}</b> lỗi: {e}")
            global _fetcher
            _fetcher = None  # reset để login lại lần sau nếu lỗi auth

        # Delay giữa các keyword khác nhau
        if kw_idx < len(keywords):
            await asyncio.sleep(5)

    return tweet_filter.filter(raw)


def _create_scan_excel(tweets: list[dict], keywords: list[str]) -> str:
    """Tạo file Excel từ danh sách tweet, trả về đường dẫn file tạm."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kết quả Scan"

    headers = [
        "STT", "Thời gian đăng", "@Username", "Tên KOL", "Followers",
        "Lượt xem", "Likes", "Retweets", "Replies", "Nội dung tweet",
        "Link bài viết", "Từ khóa khớp",
    ]
    header_fill = PatternFill(start_color="1DA1F2", end_color="1DA1F2", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alt_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
    num_fmt = "#,##0"

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, tweet in enumerate(tweets, 2):
        try:
            dt = datetime.strptime(tweet["created_at"], "%a %b %d %H:%M:%S %z %Y")
            dt_str = dt.strftime("%d/%m/%Y %H:%M")
        except Exception:
            dt_str = tweet.get("created_at", "")

        row_data = [
            row_idx - 1,
            dt_str,
            f"@{tweet.get('username', '')}",
            tweet.get("name", ""),
            tweet.get("followers", 0),
            tweet.get("views", 0),
            tweet.get("likes", 0),
            tweet.get("retweets", 0),
            tweet.get("replies", 0),
            (tweet.get("text") or "")[:150],
            tweet.get("url", ""),
            tweet.get("matched_keywords", ""),
        ]

        use_alt = (row_idx % 2 == 0)
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            if use_alt:
                cell.fill = alt_fill
            if col in (5, 6, 7, 8, 9):
                cell.number_format = num_fmt

        url = tweet.get("url", "")
        if url:
            link_cell = ws.cell(row=row_idx, column=11)
            link_cell.hyperlink = url
            link_cell.font = Font(color="0563C1", underline="single",
                                  italic=(use_alt))

    ws.freeze_panes = "A2"

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    today = datetime.now().strftime("%Y-%m-%d")
    kw_slug = "_".join(k.replace("$", "") for k in keywords[:2])
    path = os.path.join(tempfile.gettempdir(), f"scan_{kw_slug}_{today}.xlsx")
    wb.save(path)
    return path


def _format_scan_message(tweets: list[dict], keywords: list[str]) -> str:
    """Tạo message HTML tóm tắt kết quả scan."""
    kw_str = ", ".join(keywords)
    if not tweets:
        return (
            "🔍 <b>Kết quả Scan KOL</b>\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            f"❌ Không tìm thấy bài viết nào khớp tiêu chí trong {SCAN_HOURS}h qua.\n"
            f"Keywords: {kw_str}"
        )

    lines = [
        "🔍 <b>Kết quả Scan KOL</b>",
        "━━━━━━━━━━━━━━━━━━━━",
        f"📅 {SCAN_HOURS}h gần nhất",
        f"🔑 Keywords: {kw_str}",
        f"👁 Min views: {SCAN_MIN_VIEWS:,}",
        "━━━━━━━━━━━━━━━━━━━━",
        f"✅ Tìm thấy {len(tweets)} bài viết",
        "",
    ]
    for i, t in enumerate(tweets[:10], 1):
        excerpt = (t.get("text") or "")[:80].replace("<", "&lt;").replace(">", "&gt;")
        lines += [
            f"{i}. @{t.get('username', '')}",
            f"   👁 {t.get('views', 0):,} views · ❤️ {t.get('likes', 0):,} · 🔁 {t.get('retweets', 0):,}",
            f'   📝 "{excerpt}..."',
            f"   🔗 {t.get('url', '')}",
            "",
        ]
    if len(tweets) > 10:
        lines.append(f"📎 File Excel đính kèm gồm tất cả {len(tweets)} bài")
    return "\n".join(lines)


async def _send_scan_results(send_message_fn, send_document_fn, tweets: list[dict], keywords: list[str]):
    """Gửi message tóm tắt và file Excel (dùng chung cho /scan và auto scan)."""
    msg = _format_scan_message(tweets, keywords)
    await send_message_fn(msg)

    if not tweets:
        return

    try:
        excel_path = _create_scan_excel(tweets, keywords)
        kw_str = ", ".join(keywords)
        with open(excel_path, "rb") as f:
            await send_document_fn(f, os.path.basename(excel_path),
                                   f"📊 {len(tweets)} bài | Keywords: {kw_str}")
        os.remove(excel_path)
    except Exception as e:
        logger.error(f"Tạo Excel scan thất bại: {e}")


async def cmd_scan(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update.effective_user.id):
        return

    keywords = list(context.args) if context.args else CONFIG["keywords"]
    kw_str = ", ".join(keywords)
    status_msg = await update.message.reply_text(
        f"⏳ Bắt đầu quét X/Twitter...\n🔑 Keywords: <b>{kw_str}</b>\n\n"
        f"Quá trình có thể mất vài phút.",
        parse_mode="HTML",
    )

    last_text = [""]

    async def progress_cb(msg: str):
        if msg != last_text[0]:
            last_text[0] = msg
            try:
                await status_msg.edit_text(msg, parse_mode="HTML")
            except Exception:
                pass

    try:
        tweets = await _scan_kol_tweets(keywords, progress_cb=progress_cb)
    except Exception as e:
        await status_msg.edit_text(f"❌ Lỗi khi quét:\n<code>{e}</code>", parse_mode="HTML")
        return

    await status_msg.delete()

    async def send_msg(text):
        await update.message.reply_text(text, parse_mode="HTML", disable_web_page_preview=True)

    async def send_doc(f, filename, caption):
        await update.message.reply_document(document=f, filename=filename, caption=caption)

    await _send_scan_results(send_msg, send_doc, tweets, keywords)
    await update.message.reply_text("Bạn muốn làm gì tiếp theo?", reply_markup=_main_keyboard())


async def auto_scan_job(bot):
    """Chạy tự động mỗi 24h, gửi kết quả vào chat của TELEGRAM_ADMIN_ID."""
    if not TELEGRAM_ADMIN_ID:
        return

    keywords = CONFIG["keywords"]
    try:
        tweets = await _scan_kol_tweets(keywords)
    except Exception as e:
        logger.error(f"Auto scan thất bại: {e}")
        await bot.send_message(TELEGRAM_ADMIN_ID,
                               f"❌ Auto scan lỗi: <code>{e}</code>", parse_mode="HTML")
        return

    async def send_msg(text):
        await bot.send_message(TELEGRAM_ADMIN_ID, text,
                               parse_mode="HTML", disable_web_page_preview=True)

    async def send_doc(f, filename, caption):
        await bot.send_document(TELEGRAM_ADMIN_ID, document=f,
                                filename=filename, caption=caption)

    await _send_scan_results(send_msg, send_doc, tweets, keywords)


# ── Main ──────────────────────────────────────────────────────────────────────

async def _post_init(application: Application) -> None:
    await application.bot.set_my_commands([
        BotCommand("start", "Bắt đầu / xem menu chính"),
        BotCommand("monitor_x", "Kiểm tra KOL X/Twitter theo danh sách"),
        BotCommand("monitor_tg", "Kiểm tra KOL Telegram theo danh sách"),
        BotCommand("scan", "Tìm KOL mới đang đăng về dự án (24h, view >1k)"),
        BotCommand("help", "Hướng dẫn sử dụng"),
    ])
    await application.bot.set_my_description(
        "👋 Chào mừng đến với KOL Monitor Bot!\n\n"
        "Bot giúp bạn theo dõi các KOL trên X/Twitter đang đăng về dự án NEXI/Nexira.\n\n"
        "Nhấn Start để bắt đầu 👇"
    )
    await application.bot.set_my_short_description(
        "Theo dõi KOL X/Twitter đang đăng về dự án NEXI/Nexira"
    )

    scheduler = AsyncIOScheduler(timezone="Asia/Ho_Chi_Minh")
    scheduler.add_job(
        auto_scan_job,
        trigger="cron",
        hour=9,
        minute=0,
        args=[application.bot],
    )
    scheduler.start()
    logger.info("Scheduler đã khởi động (auto scan lúc 09:00 Asia/Ho_Chi_Minh)")


def main():
    app = Application.builder().token(TELEGRAM_TOKEN).post_init(_post_init).build()

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("help", cmd_help))
    app.add_handler(CommandHandler("myid", cmd_myid))
    app.add_handler(CommandHandler("adduser", cmd_adduser))
    app.add_handler(CommandHandler("removeuser", cmd_removeuser))
    app.add_handler(CommandHandler("listusers", cmd_listusers))
    app.add_handler(CommandHandler("monitor", cmd_monitor))
    app.add_handler(CommandHandler("monitor_x", cmd_monitor_x))
    app.add_handler(CommandHandler("monitor_tg", cmd_monitor_tg))
    app.add_handler(CommandHandler("scan", cmd_scan))
    app.add_handler(CallbackQueryHandler(cmd_button))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    logger.info("Bot đang chạy...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
